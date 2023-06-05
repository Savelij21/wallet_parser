import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font
from etherscan.classes.wallet_classes_v2 import Wallet, Ticker


class ExcelWallet:
    def __init__(self, wallet: Wallet):
        self.book = openpyxl.Workbook()
        self.sheet = self.book['Sheet']
        self.wallet = wallet
        self.__preparing_template()
        self.__filling_tickers_to_excel()

    def __preparing_template(self):
        self.sheet['A1'] = 'Кошелек'
        if self.wallet.show_details:
            self.sheet['B1'] = self.wallet.wallet_address
        else:
            self.sheet['C1'] = self.wallet.wallet_address

        self.sheet['D1'] = 'WinRate R'
        self.sheet['D2'] = 'WinRate UR'

        self.sheet['F1'] = 'PnL R'
        self.sheet['F2'] = 'PnL UR'

        self.sheet['H1'] = 'Average R'
        self.sheet['H2'] = 'Average UR'

        if self.wallet.show_details:
            self.sheet['J1'] = 'Total R'
            self.sheet['J2'] = 'Total UR'
        else:
            self.sheet['L1'] = 'Total R'
            self.sheet['L2'] = 'Total UR'

        # ticker params and metrics
        table_headers = [
            'Тикеры',
            'Дата (YYYY-MM-DD), время',
            'Всего купили',
            'Всего продали',
            'Дельта',
            'Дельта, $',
            'Сумма покупок, $',
            'Сумма продаж, $',
            'Дельта (профит), $',
            'Цена покупки, $',
            'Цена продажи, $',
            'Текущая цена, $',
            'Покупок',
            'Продаж',
            'Объем торгов за сутки, $'
        ]
        self.sheet['A4'] = table_headers[0]
        self.sheet['B4'] = table_headers[1]
        self.sheet['C4'] = table_headers[2]
        self.sheet['D4'] = table_headers[3]
        self.sheet['E4'] = table_headers[4]
        self.sheet['F4'] = table_headers[5]
        self.sheet['G4'] = table_headers[6]
        self.sheet['H4'] = table_headers[7]
        self.sheet['I4'] = table_headers[8]
        self.sheet['J4'] = table_headers[9]
        self.sheet['K4'] = table_headers[10]
        self.sheet['L4'] = table_headers[11]
        self.sheet['M4'] = table_headers[12]
        self.sheet['N4'] = table_headers[13]
        self.sheet['O4'] = table_headers[14]

        if self.wallet.show_details:
            self.sheet.column_dimensions['B'].width = 20
            self.sheet.column_dimensions['J'].width = 15
            self.sheet.column_dimensions['K'].width = 15
            self.sheet.column_dimensions['L'].width = 15
        else:
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['J'].width = 10
            self.sheet.column_dimensions['K'].width = 10
            self.sheet.column_dimensions['L'].width = 24

        # edit columns width
        self.sheet.column_dimensions['A'].width = 12
        self.sheet.column_dimensions['C'].width = 15
        self.sheet.column_dimensions['D'].width = 15
        self.sheet.column_dimensions['E'].width = 15
        self.sheet.column_dimensions['F'].width = 15
        self.sheet.column_dimensions['G'].width = 15
        self.sheet.column_dimensions['H'].width = 15
        self.sheet.column_dimensions['I'].width = 15
        self.sheet.column_dimensions['M'].width = 10
        self.sheet.column_dimensions['N'].width = 10
        self.sheet.column_dimensions['O'].width = 24

        # borders for header string
        header_string_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        for i in range(1, 16):
            self.sheet.cell(row=4, column=i).border = header_string_border

        # freezing header rows and columns
        self.sheet.freeze_panes = self.sheet['B5']

    def __filling_tickers_to_excel(self):
        # styles
        left_border = Border(left=Side(style='medium'))
        right_border = Border(right=Side(style='medium'))
        bottom_border = Border(bottom=Side(style='medium'))

        grey_fill = PatternFill('solid', fgColor='A0A0A0')
        blue_fill = PatternFill('solid', fgColor='3399FF')
        red_fill = PatternFill('solid', fgColor='FF6666')
        green_fill = PatternFill('solid', fgColor='99FF99')
        yellow_fill = PatternFill('solid', fgColor='FFFF66')
        black_fill = PatternFill('solid', fgColor='606060')

        # filling data
        def fill_total_string_for_ticker(ticker_obj: Ticker, string_num: int):  # only data, no styles
            # TOTAL DATA FOR TICKER
            string = str(string_num)

            # token symbol column
            self.sheet[f'A{string}'] = f'{ticker_obj.ticker_symbol}                      {ticker_obj.token_address}'
            if not ticker.is_valid or round(ticker.total_delta_usd, 1) < 0:
                self.sheet[f'A{string}'].fill = red_fill
            elif ticker.is_stablecoin:
                self.sheet[f'A{string}'].fill = blue_fill

            # datetime field skip
            self.sheet[f'B{string}'] = '---'

            # tools
            def validated(param, round_num=2):
                if param == 'N/A':
                    return 'N/A'
                else:
                    return round(float(param), round_num)

            # total params
            self.sheet[f'C{string}'] = ticker_obj.total_bought
            self.sheet[f'D{string}'] = ticker_obj.total_sold
            self.sheet[f'E{string}'] = ticker_obj.total_delta
            self.sheet[f'F{string}'] = validated(ticker_obj.total_delta_usd, 2)
            # sum metrics
            self.sheet[f'G{string}'] = validated(ticker_obj.sum_bought_usd, 2)
            self.sheet[f'H{string}'] = validated(ticker_obj.sum_sold_usd, 2)

            self.sheet[f'I{string}'] = validated(ticker_obj.sum_delta_usd, 2)
            self.sheet[f'I{string}'].font = Font(color='000000')
            if ticker_obj.sum_delta_usd == 'N/A':
                self.sheet[f'I{string}'].fill = yellow_fill
            elif ticker_obj.sum_delta_usd <= 0:
                self.sheet[f'I{string}'].fill = red_fill
            else:
                self.sheet[f'I{string}'].fill = green_fill

            # current price
            self.sheet[f'L{string}'] = ticker_obj.current_price_usd

            # transactions amount
            self.sheet[f'M{string}'] = ticker_obj.buy_tr_amount
            self.sheet[f'N{string}'] = ticker_obj.sell_tr_amount

            # volume
            self.sheet[f'O{string}'] = ticker_obj.volume_24h_usd

            # borders for total string
            self.sheet[f'A{string}'].border = left_border
            self.sheet[f'O{string}'].border = right_border
            for i in range(1, 16):
                self.sheet.cell(row=current_string, column=i).border = bottom_border

        current_string = 5  # first string after header
        if self.wallet.show_details:  # if user wants to get transactions details
            for ticker in self.wallet.tickers:
                # skip stables
                if ticker.is_stablecoin:
                    continue
                # FOR SINGLE TRANSACTION
                for transaction in ticker.transactions:
                    string = str(current_string)
                    # token symbol column
                    self.sheet[f'A{string}'] = transaction.token_symbol
                    if not ticker.is_valid:
                        self.sheet[f'A{string}'].fill = red_fill
                    # transaction data column
                    self.sheet[f'B{string}'] = transaction.date_time
                    # fields depending on buying/selling type
                    if transaction.type == 'buying':
                        self.sheet[f'C{string}'] = transaction.value
                        self.sheet[f'G{string}'] = transaction.sum_usd
                        self.sheet[f'J{string}'] = transaction.token_price_usd
                    else:
                        self.sheet[f'D{string}'] = transaction.value
                        self.sheet[f'H{string}'] = transaction.sum_usd
                        self.sheet[f'K{string}'] = transaction.token_price_usd
                    # borders for string with transaction
                    self.sheet[f'A{string}'].border = left_border
                    self.sheet[f'O{string}'].border = right_border
                    # string for next transaction or for total info of ticker
                    current_string += 1

                # filling data for total ticker string
                fill_total_string_for_ticker(ticker_obj=ticker, string_num=current_string)

                # styles for total string of detailed ticker
                string = str(current_string)
                # painting token symbol column
                if ticker.is_valid and round(ticker.total_delta_usd, 1) >= 0 and not ticker.is_stablecoin:
                    self.sheet[f'A{string}'].fill = grey_fill

                # painting fields to grey if token is not scammed
                grey_fields = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'N', 'O']
                for column in grey_fields:
                    self.sheet[f'{column}{string}'].fill = grey_fill

                # coloring string to black if token is SCAMMED
                if ticker.is_scam:
                    for i in range(1, 16):
                        if i == 9:  # skip sum_delta column to not paint this field to black
                            continue
                        self.sheet.cell(row=current_string, column=i).font = Font(color='FFFFFF')
                        self.sheet.cell(row=current_string, column=i).fill = black_fill

                # adding string for the next ticker
                current_string += 1

        else:  # if user does not want to get transactions

            for ticker in self.wallet.tickers:
                # skip stables
                if ticker.is_stablecoin:
                    continue
                # TOTAL DATA FOR TICKER
                # coloring string if token is SCAMMED (before data filling because usual data not need in styles)
                if ticker.is_scam:
                    for i in range(1, 16):
                        if i == 9:  # skip sum_delta column to not paint this field to black
                            continue
                        self.sheet.cell(row=current_string, column=i).font = Font(color='FFFFFF')
                        self.sheet.cell(row=current_string, column=i).fill = black_fill
                # filling data
                fill_total_string_for_ticker(ticker_obj=ticker, string_num=current_string)
                # styles for data, that is different from detailed tickers
                string = str(current_string)

                # adding string for the next ticker
                current_string += 1

        if not self.wallet.show_details:
            self.sheet.delete_cols(2)
            self.sheet.delete_cols(9)
            self.sheet.delete_cols(9)


    def save(self, file_name: str):
        self.book.save(file_name)


