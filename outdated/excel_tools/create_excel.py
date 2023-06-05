import openpyxl
from openpyxl.styles import PatternFill, Border, Side


def create_template(sheet):

    sheet['A1'] = 'Кошелек'

    sheet['D1'] = 'WinRate R'
    sheet['D2'] = 'WinRate UR'

    sheet['F1'] = 'PnL R'
    sheet['F2'] = 'PnL UR'

    sheet['H1'] = 'Average R'
    sheet['H2'] = 'Average UR'

    sheet['J1'] = 'Total R'
    sheet['J2'] = 'Total UR'

    # ticker params and metrics
    table_headers = [
        'Тикеры',
        'Дата, время',
        'Всего купили',
        'Всего продали',
        'Дельта',
        'Дельта, $',
        'Сумма покупок, $',
        'Сумма продаж, $',
        'Дельта',
        'Цена покупки, $',
        'Цена продажи, $',
        'Текущая цена, $',
        'Покупок',
        'Продаж',
    ]
    sheet['A4'] = table_headers[0]
    sheet['B4'] = table_headers[1]
    sheet['C4'] = table_headers[2]
    sheet['D4'] = table_headers[3]
    sheet['E4'] = table_headers[4]
    sheet['F4'] = table_headers[5]
    sheet['G4'] = table_headers[6]
    sheet['H4'] = table_headers[7]
    sheet['I4'] = table_headers[8]
    sheet['J4'] = table_headers[9]
    sheet['K4'] = table_headers[10]
    sheet['L4'] = table_headers[11]
    sheet['M4'] = table_headers[12]
    sheet['N4'] = table_headers[13]

    # edit columns width
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['K'].width = 15
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['M'].width = 15
    sheet.column_dimensions['N'].width = 15

    # borders
    # borders for header string
    header_string_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    for i in range(1, 15):
        sheet.cell(row=4, column=i).border = header_string_border


def ticker_block(ticker, current_string, sheet):
    ticker_transactions = ticker['transactions']
    ticker_transactions.reverse()
    ticker_name = ticker_transactions[0]['tokenSymbol']

    # borders
    left_border = Border(left=Side(style='medium'))
    right_border = Border(right=Side(style='medium'))
    bottom_border = Border(bottom=Side(style='medium'))

    for transaction in ticker_transactions:
        sheet[f'A{str(current_string)}'] = transaction['tokenSymbol']
        sheet[f'B{str(current_string)}'] = transaction['date_time']

        if transaction['type'] == 'buying':
            sheet[f'C{str(current_string)}'] = round(transaction['value'], 3)  # buying tokens amount
            sheet[f'G{str(current_string)}'] = round(transaction['sum_usd'], 3)  # buying sum
            sheet[f'J{str(current_string)}'] = round(transaction['token_price_usd'], 6)  # buying price
        else:
            sheet[f'D{str(current_string)}'] = round(transaction['value'], 3)  # selling tokens amount
            sheet[f'H{str(current_string)}'] = round(transaction['sum_usd'], 3)  # buying sum
            sheet[f'K{str(current_string)}'] = round(transaction['token_price_usd'], 6)  # selling price

        sheet[f'A{str(current_string)}'].border = left_border
        sheet[f'N{str(current_string)}'].border = right_border

        current_string += 1

    # styles
    grey_fill = PatternFill('solid', fgColor='A0A0A0')
    red_fill = PatternFill('solid', fgColor='FF6666')
    green_fill = PatternFill('solid', fgColor='99FF99')
    yellow_fill = PatternFill('solid', fgColor='FFFF66')

    # total metrics
    sheet[f'A{str(current_string)}'] = ticker_name
    if round(ticker['total_delta'], 1) < 0:  # if error in data from api
        sheet[f'A{str(current_string)}'].fill = red_fill
    elif ticker['current_price_usd'] == 'N/A' or ticker['sum_bought_usd'] == 'N/A' or ticker['sum_sold_usd'] == 'N/A':
        sheet[f'A{str(current_string)}'].fill = yellow_fill
    else:
        sheet[f'A{str(current_string)}'].fill = grey_fill

    # test
    def validated(param, round_num=3):
        if param == 'N/A':
            return 'N/A'
        else:
            return round(float(param), round_num)

    # test
    sheet[f'B{str(current_string)}'].fill = grey_fill

    sheet[f'C{str(current_string)}'] = validated(ticker['total_bought'], 3)
    sheet[f'C{str(current_string)}'].fill = grey_fill

    sheet[f'D{str(current_string)}'] = validated(ticker['total_sold'], 3)
    sheet[f'D{str(current_string)}'].fill = grey_fill

    sheet[f'E{str(current_string)}'] = validated(ticker['total_delta'], 3)
    sheet[f'E{str(current_string)}'].fill = grey_fill

    sheet[f'F{str(current_string)}'] = validated(ticker['total_delta_usd'], 3)
    sheet[f'F{str(current_string)}'].fill = grey_fill

    # sum metrics
    sheet[f'G{str(current_string)}'] = validated(ticker['sum_bought_usd'], 3)
    sheet[f'G{str(current_string)}'].fill = grey_fill

    sheet[f'H{str(current_string)}'] = validated(ticker['sum_sold_usd'], 3)
    sheet[f'H{str(current_string)}'].fill = grey_fill

    sheet[f'I{str(current_string)}'] = validated(ticker['sum_delta_usd'], 3)

    if ticker['sum_delta_usd'] != 'N/A' and float(ticker['sum_delta_usd']) >= 0:
        sheet[f'I{str(current_string)}'].fill = green_fill
    else:
        sheet[f'I{str(current_string)}'].fill = red_fill

    # painting empty cells
    sheet[f'J{str(current_string)}'].fill = grey_fill
    sheet[f'K{str(current_string)}'].fill = grey_fill

    # current price
    sheet[f'L{str(current_string)}'] = validated(ticker['current_price_usd'], 6)
    sheet[f'L{str(current_string)}'].fill = grey_fill

    # transactions amount
    sheet[f'M{str(current_string)}'] = ticker['buying_amount']
    sheet[f'M{str(current_string)}'].fill = grey_fill

    sheet[f'N{str(current_string)}'] = ticker['selling_amount']
    sheet[f'N{str(current_string)}'].fill = grey_fill

    # borders for total string
    sheet[f'A{str(current_string)}'].border = left_border
    sheet[f'N{str(current_string)}'].border = right_border

    for i in range(1, 15):
        sheet.cell(row=current_string, column=i).border = bottom_border


